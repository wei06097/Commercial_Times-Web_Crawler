import os
import openpyxl
FILE_NAME = "ctee_news.xlsx"

### 沒有檔案就建立
def check_xlsx():
    if os.path.exists(FILE_NAME): return
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'id'
    sheet['B1'] = 'Title'
    sheet['C1'] = 'Link'
    sheet['D1'] = 'Datetime'
    sheet['E1'] = 'Author'
    sheet['F1'] = 'Content'
    workbook.save(FILE_NAME)

### 檢查資料是否已存在
def check_existance(id):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active
    existing = False
    column_to_search = sheet['A']
    for cell in column_to_search[::-1]:
        if cell.value == id:
            existing = True
            break
    return existing

### 新增一筆資料
def add_data(data):
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active
    row = sheet.max_row + 1
    for col, value in enumerate(data, start=1):
        cell = sheet.cell(row=row, column=col, value=value)
    workbook.save(FILE_NAME)
